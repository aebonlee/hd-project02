# -*- coding: utf-8 -*-
"""
과제 A. 업체별 상이한 PDF에서 데이터 추출 → 엑셀

실행:
    python3 taskA_pdf_extract/extract.py                # sample_pdfs/ 전체 처리
    python3 taskA_pdf_extract/extract.py 경로1.pdf ...  # 특정 파일만 처리

동작:
    1. pdfplumber 로 PDF 텍스트 추출
    2. vendors.json 의 키워드로 업체 판별
    3. 공통 4개 항목(품번/무게/수량/금액)은 공통 정규식 → 실패 시 값 형식 단독 패턴(fallback)
    4. 인코텀즈 번호는 업체별 패턴으로 분기 추출
    5. 추출 실패 항목은 빈칸 + '검토필요' 플래그 → output/추출결과.xlsx 저장

새 업체 추가: vendors.json 의 vendors 에 키워드/인코텀즈 패턴만 추가하면 된다.
(선택) 정규식으로 실패한 인코텀즈는 LLM API 보조 추출로 확장 가능 — README 참조.
"""
import json
import os
import re
import sys

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, "vendors.json")
SAMPLE_DIR = os.path.join(BASE_DIR, "sample_pdfs")
OUT_PATH = os.path.join(BASE_DIR, "output", "추출결과.xlsx")

COLUMNS = ["파일명", "업체명", "품번", "무게", "수량", "금액", "인코텀즈", "검토필요"]


def load_config():
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)


def read_pdf_text(path):
    """PDF 전체 페이지의 텍스트를 하나의 문자열로 합친다."""
    chunks = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            chunks.append(text)
            # 표 형식 문서 대비: 표 셀 텍스트도 함께 수집
            for table in page.extract_tables():
                for row in table:
                    chunks.append(" ".join(c or "" for c in row))
    return "\n".join(chunks)


def detect_vendor(text, filename, vendors_cfg):
    """본문 또는 파일명에서 업체 키워드를 찾아 업체명을 판별한다."""
    for vendor_name, rule in vendors_cfg.items():
        for kw in rule["키워드"]:
            if kw in text or kw in filename:
                return vendor_name
    return None


def extract_field(text, primary_pattern, fallback_pattern=None):
    """공통 패턴 → 실패 시 값 형식 단독 패턴 순으로 시도."""
    m = re.search(primary_pattern, text)
    if m:
        return m.group(1)
    if fallback_pattern:
        m = re.search(fallback_pattern, text)
        if m:
            return m.group(1)
    return ""


def extract_from_pdf(path, config):
    filename = os.path.basename(path)
    text = read_pdf_text(path)

    vendor = detect_vendor(text, filename, config["vendors"])
    row = {"파일명": filename, "업체명": vendor or "", "검토필요": ""}
    missing = []

    # 공통 4개 항목
    for field, rule in config["common_fields"].items():
        fallback = config.get("fallback_fields", {}).get(field)
        value = extract_field(text, rule["패턴"], fallback)
        row[field] = value
        if not value:
            missing.append(field)

    # 인코텀즈 — 업체별 규칙 분기
    incoterms = ""
    if vendor:
        incoterms = extract_field(text, config["vendors"][vendor]["인코텀즈_패턴"])
    row["인코텀즈"] = incoterms
    if not incoterms:
        missing.append("인코텀즈")
    if not vendor:
        missing.insert(0, "업체판별")

    row["검토필요"] = ("Y (" + ", ".join(missing) + ")") if missing else "N"
    return row


def save_excel(rows, out_path):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "추출결과"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    warn_fill = PatternFill("solid", fgColor="FDECEC")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row in rows:
        ws.append([row.get(col, "") for col in COLUMNS])
        if row["검토필요"].startswith("Y"):
            for cell in ws[ws.max_row]:
                cell.fill = warn_fill

    widths = [32, 12, 14, 11, 9, 12, 16, 22]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"
    wb.save(out_path)


def main():
    config = load_config()

    if len(sys.argv) > 1:
        pdf_paths = sys.argv[1:]
    else:
        pdf_paths = sorted(
            os.path.join(SAMPLE_DIR, f)
            for f in os.listdir(SAMPLE_DIR)
            if f.lower().endswith(".pdf")
        )
    if not pdf_paths:
        print("처리할 PDF가 없습니다. 먼저 make_samples.py 를 실행하세요.")
        sys.exit(1)

    rows = []
    for path in pdf_paths:
        row = extract_from_pdf(path, config)
        rows.append(row)
        status = "OK" if row["검토필요"] == "N" else row["검토필요"]
        print(f"[추출] {row['파일명']} → 업체={row['업체명'] or '?'} "
              f"품번={row['품번']} 무게={row['무게']}kg 수량={row['수량']} "
              f"금액=USD {row['금액']} 인코텀즈={row['인코텀즈']} / {status}")

    save_excel(rows, OUT_PATH)
    print(f"\n[저장] {OUT_PATH} ({len(rows)}건)")


if __name__ == "__main__":
    main()
