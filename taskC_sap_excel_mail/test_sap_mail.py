# -*- coding: utf-8 -*-
"""과제 C 순수 로직 테스트 (색 변환 · 메일 본문 · 범위 → HTML 표)

실행:
    python3 taskC_sap_excel_mail/test_sap_mail.py

색 변환/본문 테스트는 표준 라이브러리만으로 동작하고,
범위 → HTML 표 테스트는 openpyxl 이 있을 때만 실행된다.
"""
import os
import sys
import unittest

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import sap_mail  # noqa: E402

try:
    import openpyxl  # noqa: F401
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False


class FakeColor:
    """openpyxl 색 객체의 type/rgb 속성만 흉내내는 테스트용 객체."""

    def __init__(self, type_, rgb):
        self.type = type_
        self.rgb = rgb


class TestColorOf(unittest.TestCase):
    def test_rgb_aarrggbb(self):
        self.assertEqual(sap_mail._color_of(FakeColor("rgb", "FF1F4E79")), "#1F4E79")

    def test_rgb_rrggbb(self):
        self.assertEqual(sap_mail._color_of(FakeColor("rgb", "D64545")), "#D64545")

    def test_none_and_theme(self):
        self.assertIsNone(sap_mail._color_of(None))
        self.assertIsNone(sap_mail._color_of(FakeColor("theme", None)))
        self.assertIsNone(sap_mail._color_of(FakeColor("rgb", None)))


class TestBuildMailHtml(unittest.TestCase):
    def test_body_contains_table_and_date(self):
        html = sap_mail.build_mail_html("<table><tr><td>표</td></tr></table>", "2026-08-22")
        self.assertIn("2026-08-22", html)
        self.assertIn("<table>", html)
        self.assertIn("자재 수급 현황", html)
        self.assertIn("첨부 엑셀", html)


@unittest.skipUnless(HAS_OPENPYXL, "openpyxl 미설치 — 범위 → HTML 표 테스트 생략")
class TestRangeToHtml(unittest.TestCase):
    def _make_sheet(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "구분"
        ws["B1"] = "수량"
        ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
        ws["A1"].font = Font(color="FFFFFFFF", bold=True)
        ws["A2"] = "볼트"
        ws["B2"] = 1200
        ws.merge_cells("A3:B3")
        ws["A3"] = "합계"
        return ws

    def test_styles_numbers_and_merge(self):
        ws = self._make_sheet()
        html = sap_mail.range_to_html(ws, "A1:B3")
        self.assertIn("background:#1F4E79", html)   # 배경색 유지
        self.assertIn("font-weight:bold", html)      # 굵기 유지
        self.assertIn("1,200", html)                 # 숫자 콤마 서식
        self.assertIn("text-align:right", html)      # 숫자 우측 정렬
        self.assertIn('colspan="2"', html)           # 병합 셀
        self.assertEqual(html.count("<tr>"), 3)

    def test_sample_excel_if_present(self):
        if not os.path.exists(sap_mail.SRC_XLSX):
            self.skipTest("샘플 엑셀 없음 (make_sample.py 미실행)")
        from openpyxl import load_workbook
        ws = load_workbook(sap_mail.SRC_XLSX)[sap_mail.CAPTURE_SHEET]
        html = sap_mail.range_to_html(ws, sap_mail.CAPTURE_RANGE)
        self.assertIn("<table", html)
        self.assertIn("<td", html)


if __name__ == "__main__":
    unittest.main(verbosity=2)
