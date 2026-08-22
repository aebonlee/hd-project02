# -*- coding: utf-8 -*-
"""
과제 C. SAP 다운로드 → 엑셀 가공 → 팀 메일 송부 (가공 이후 단계 자동화)

실행:
    python3 taskC_sap_excel_mail/sap_mail.py

동작:
    1. 가공 완료된 엑셀(sample_data/자재수급현황_가공완료.xlsx)을 연다
       — 실제 업무에서는 팀즈 마스터 쿼리 파일로 가공된 결과 파일 경로를 지정
    2. 지정 범위(요약!A1:G10)를 셀 서식(배경색/글자색/굵기)을 살린 HTML 표로 변환해
       메일 본문에 삽입 — 개인 환경에서는 '이미지 캡처' 대신 HTML 표로 대체
       (xlwings/PIL 이미지 캡처 방법은 README 참조)
    3. 엑셀을 날짜가 붙은 다른 이름(자재수급현황_YYYYMMDD.xlsx)으로 저장
    4. 본문 + 첨부가 포함된 .eml 메일 초안 생성
       — 회사 PC에서는 Outlook COM 발송으로 교체 (README)
"""
import os
import shutil
from datetime import date
from email.message import EmailMessage
from email.utils import formatdate

# openpyxl 은 실제로 필요한 함수 안에서 import 한다.
# (라이브러리가 없는 환경에서도 순수 로직 함수를 import 할 수 있도록)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SRC_XLSX = os.path.join(BASE_DIR, "sample_data", "자재수급현황_가공완료.xlsx")
OUT_DIR = os.path.join(BASE_DIR, "output")

CAPTURE_SHEET = "요약"       # 메일 본문에 넣을 시트
CAPTURE_RANGE = "A1:G10"    # 메일 본문에 넣을 범위 (회사 파일에 맞게 수정)

MAIL_TO = ["team_member1@example.com", "team_member2@example.com"]
MAIL_CC = ["team_leader@example.com"]
MAIL_FROM = "material_report@example.com"


def _color_of(color_obj):
    """openpyxl 색 객체 → CSS hex (#RRGGBB). 테마색 등 미해석 색은 None."""
    if color_obj is None or color_obj.type != "rgb" or not color_obj.rgb:
        return None
    rgb = str(color_obj.rgb)
    if len(rgb) == 8:  # AARRGGBB
        rgb = rgb[2:]
    if rgb == "000000" and color_obj.type != "rgb":
        return None
    return f"#{rgb}"


def range_to_html(ws, cell_range):
    """셀 범위를 배경색/글자색/굵기/정렬을 반영한 HTML 표로 변환한다.

    [대체 구현 메모] 기획서의 '특정 범위 이미지 캡처'는 엑셀 설치가 필요한
    xlwings COM 기능이라 개인 환경에서는 HTML 표로 대체한다. (README 참조)
    """
    from openpyxl.utils import range_boundaries

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    merged = {}
    skip = set()
    for mc in ws.merged_cells.ranges:
        merged[(mc.min_row, mc.min_col)] = (mc.max_row - mc.min_row + 1,
                                            mc.max_col - mc.min_col + 1)
        for r in range(mc.min_row, mc.max_row + 1):
            for c in range(mc.min_col, mc.max_col + 1):
                if (r, c) != (mc.min_row, mc.min_col):
                    skip.add((r, c))

    rows_html = []
    for r in range(min_row, max_row + 1):
        cells = []
        for c in range(min_col, max_col + 1):
            if (r, c) in skip:
                continue
            cell = ws.cell(row=r, column=c)
            value = "" if cell.value is None else cell.value
            style = ["border:1px solid #b9c6d6", "padding:5px 9px",
                     "font-size:13px", "font-family:'Malgun Gothic',sans-serif"]
            if cell.fill and cell.fill.patternType == "solid":
                bg = _color_of(cell.fill.fgColor)
                if bg:
                    style.append(f"background:{bg}")
            if cell.font:
                if cell.font.bold:
                    style.append("font-weight:bold")
                fg = _color_of(cell.font.color)
                if fg and fg != "#000000":
                    style.append(f"color:{fg}")
            if cell.alignment and cell.alignment.horizontal:
                style.append(f"text-align:{cell.alignment.horizontal}")
            elif isinstance(value, (int, float)):
                style.append("text-align:right")
            span = ""
            if (r, c) in merged:
                rs, cs = merged[(r, c)]
                if rs > 1:
                    span += f' rowspan="{rs}"'
                if cs > 1:
                    span += f' colspan="{cs}"'
            if isinstance(value, (int, float)):
                value = f"{value:,}"
            cells.append(f'<td style="{";".join(style)}"{span}>{value}</td>')
        rows_html.append("<tr>" + "".join(cells) + "</tr>")
    return '<table style="border-collapse:collapse;">' + "\n".join(rows_html) + "</table>"


def build_mail_html(table_html, today_str):
    return f"""<html>
<body style="font-family:'Malgun Gothic','맑은 고딕',sans-serif;color:#22303f;">
  <p style="font-size:13px;">팀원 여러분 안녕하세요.<br>
  {today_str} 일일 자재 수급 현황을 아래와 같이 공유드립니다.<br>
  상세 데이터는 첨부 엑셀 파일을 확인해 주세요.</p>
  {table_html}
  <p style="font-size:12px;color:#6b7a8c;">* 본 메일은 자동 생성되었습니다. (원본: SAP 다운로드 → 마스터 쿼리 가공)</p>
</body>
</html>"""


def main():
    from openpyxl import load_workbook

    if not os.path.exists(SRC_XLSX):
        raise SystemExit("샘플 엑셀이 없습니다. 먼저 make_sample.py 를 실행하세요.")
    os.makedirs(OUT_DIR, exist_ok=True)
    today = date.today()
    today_str = today.isoformat()
    ymd = today.strftime("%Y%m%d")

    # 1) 가공 완료 엑셀 열기
    wb = load_workbook(SRC_XLSX)
    ws = wb[CAPTURE_SHEET]

    # 2) 지정 범위 → HTML 표 (이미지 캡처 대체)
    table_html = range_to_html(ws, CAPTURE_RANGE)
    body_html = build_mail_html(table_html, today_str)

    # 3) 다른 이름(날짜)으로 저장 → 첨부 파일
    attach_name = f"자재수급현황_{ymd}.xlsx"
    attach_path = os.path.join(OUT_DIR, attach_name)
    shutil.copyfile(SRC_XLSX, attach_path)

    # 미리보기 HTML도 저장
    preview_path = os.path.join(OUT_DIR, f"sap_mail_preview_{ymd}.html")
    with open(preview_path, "w", encoding="utf-8") as f:
        f.write(body_html)

    # 4) .eml 메일 초안 생성 (회사 PC에서는 Outlook COM 발송으로 교체)
    msg = EmailMessage()
    msg["Subject"] = f"[자재] 일일 자재 수급 현황 공유 ({today_str})"
    msg["From"] = MAIL_FROM
    msg["To"] = ", ".join(MAIL_TO)
    msg["Cc"] = ", ".join(MAIL_CC)
    msg["Date"] = formatdate(localtime=True)
    msg.set_content("HTML 지원 메일 클라이언트에서 확인해 주세요.")
    msg.add_alternative(body_html, subtype="html")
    with open(attach_path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=attach_name,
        )

    eml_path = os.path.join(OUT_DIR, f"sap_mail_{ymd}.eml")
    with open(eml_path, "wb") as f:
        f.write(msg.as_bytes())

    print(f"[저장] 첨부용 엑셀   : {attach_path}")
    print(f"[저장] 본문 미리보기 : {preview_path}")
    print(f"[저장] 메일 초안     : {eml_path}")
    print(f"[정보] 본문 삽입 범위: {CAPTURE_SHEET}!{CAPTURE_RANGE} → HTML 표 변환 완료")


if __name__ == "__main__":
    main()
