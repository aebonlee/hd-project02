# -*- coding: utf-8 -*-
"""
과제 C 데모용 '가공 완료' 샘플 엑셀 생성 스크립트

실행:
    python3 taskC_sap_excel_mail/make_sample.py

실제 업무에서는 SAP 다운로드 → 팀즈 마스터 쿼리(Power Query, VLOOKUP·피벗) 가공 결과에
해당하는 파일이다. 여기서는 그 결과물 형태(요약 피벗 표)를 직접 생성한다.
"""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_PATH = os.path.join(BASE_DIR, "sample_data", "자재수급현황_가공완료.xlsx")

ROWS = [
    ["구분", "품번", "품명", "필요수량", "가용재고", "부족수량", "입고예정일"],
    ["긴급", "HD-48213-AB", "브라켓 어셈블리", 320, 180, 140, "2026-08-23"],
    ["긴급", "HD-91422-EF", "유압 실린더 로드", 480, 210, 270, "2026-08-24"],
    ["주의", "HD-77105-CD", "서포트 빔", 150, 120, 30, "2026-08-25"],
    ["주의", "HD-33518-IJ", "체결 볼트 세트", 220, 190, 30, "2026-08-25"],
    ["정상", "HD-52110-CD", "밸브 커버", 160, 240, 0, "-"],
    ["정상", "HD-64007-GH", "가스켓 씰", 90, 150, 0, "-"],
    ["정상", "HD-80331-EF", "베어링 하우징", 70, 95, 0, "-"],
]


def main():
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "요약"

    thin = Side(style="thin", color="B9C6D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    ws["A1"] = "일일 자재 수급 현황 (가공 완료)"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws.merge_cells("A1:G1")

    grade_fill = {
        "긴급": PatternFill("solid", fgColor="FDECEC"),
        "주의": PatternFill("solid", fgColor="FDF3DF"),
    }
    for r, row in enumerate(ROWS, start=3):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = border
            if r == 3:
                cell.fill = PatternFill("solid", fgColor="1F4E79")
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.alignment = center
            else:
                if row[0] in grade_fill:
                    cell.fill = grade_fill[row[0]]
                if c == 6 and isinstance(value, int) and value > 0:
                    cell.font = Font(bold=True, color="D64545")
                if c in (1, 7):
                    cell.alignment = center

    widths = [8, 14, 17, 10, 10, 10, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT_PATH)
    print(f"[생성] {OUT_PATH}")


if __name__ == "__main__":
    main()
