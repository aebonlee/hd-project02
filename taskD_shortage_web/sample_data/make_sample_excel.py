# -*- coding: utf-8 -*-
"""
결품현황 샘플 엑셀(결품현황_샘플.xlsx) + 웹앱 내장용 샘플 데이터(js/sample_data.js) 생성 스크립트

실행:
    python3 taskD_shortage_web/sample_data/make_sample_excel.py

- 5개 협력업체, 40행 이상의 샘플 데이터를 생성한다.
- 동일한 데이터를 엑셀 파일과 자바스크립트 파일 양쪽에 기록하여
  "샘플 데이터 불러오기" 버튼과 엑셀 업로드 데모가 항상 같은 결과를 내도록 한다.
"""
import json
import os
import random
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE_DIR, "결품현황_샘플.xlsx")
JS_PATH = os.path.join(os.path.dirname(BASE_DIR), "js", "sample_data.js")

HEADERS = [
    "일자", "업체코드", "업체명", "담당자명", "담당자 연락처",
    "품번", "품명", "필요수량", "확보수량", "결품수량", "확정구간", "라인",
]

VENDORS = [
    ("V001", "대한정밀", "김철수", "010-1234-1001"),
    ("V002", "한빛금속", "이영희", "010-2345-2002"),
    ("V003", "세종산업", "박민준", "010-3456-3003"),
    ("V004", "동성테크", "최수아", "010-4567-4004"),
    ("V005", "미래부품", "정재훈", "010-5678-5005"),
]

PART_NAMES = [
    "브라켓 어셈블리", "서포트 빔", "유압 실린더 로드", "체결 볼트 세트", "밸브 커버",
    "베어링 하우징", "가스켓 씰", "커넥팅 파이프", "고정 클램프", "센서 마운트",
    "스페이서 링", "플랜지 커버", "체인 가이드", "리프팅 러그", "배관 엘보",
]

LINES = ["1라인", "2라인", "3라인", "4라인"]
SECTIONS = ["D+1", "D+2", "D+3", "D+4"]


def build_rows():
    random.seed(42)
    today = date.today().isoformat()
    rows = []
    seq = 0
    # 업체별 8~10개 품목 → 총 40행 이상
    per_vendor = [9, 8, 9, 8, 10]
    for (code, name, manager, phone), n in zip(VENDORS, per_vendor):
        for i in range(n):
            seq += 1
            need = random.choice([120, 200, 240, 300, 360, 480, 600])
            secured = int(need * random.choice([0.3, 0.5, 0.6, 0.7, 0.8]))
            rows.append([
                today,
                code,
                name,
                manager,
                phone,
                f"HD-{random.randint(10000, 99999)}-{chr(65 + seq % 26)}{chr(66 + i % 24)}",
                random.choice(PART_NAMES),
                need,
                secured,
                need - secured,
                random.choice(SECTIONS),
                random.choice(LINES),
            ])
    return rows


def write_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "결품현황"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row in rows:
        ws.append(row)

    widths = [11, 9, 11, 10, 15, 15, 17, 9, 9, 9, 9, 8]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"

    wb.save(XLSX_PATH)
    print(f"[생성] {XLSX_PATH} ({len(rows)}행)")


def write_js(rows):
    records = [dict(zip(HEADERS, row)) for row in rows]
    payload = json.dumps(records, ensure_ascii=False, indent=2)
    content = (
        "// 자동 생성 파일 — sample_data/make_sample_excel.py 로 재생성\n"
        "// 결품현황_샘플.xlsx 와 동일한 데이터 (샘플 데이터 불러오기 버튼용)\n"
        "(function (global) {\n"
        "  'use strict';\n"
        f"  var SAMPLE_SHORTAGE_DATA = {payload};\n"
        "  if (typeof module !== 'undefined' && module.exports) {\n"
        "    module.exports = SAMPLE_SHORTAGE_DATA;\n"
        "  } else {\n"
        "    global.SAMPLE_SHORTAGE_DATA = SAMPLE_SHORTAGE_DATA;\n"
        "  }\n"
        "})(typeof window !== 'undefined' ? window : this);\n"
    )
    with open(JS_PATH, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"[생성] {JS_PATH} ({len(rows)}건)")


if __name__ == "__main__":
    rows = build_rows()
    write_xlsx(rows)
    write_js(rows)
